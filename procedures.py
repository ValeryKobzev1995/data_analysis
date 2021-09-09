import xlrd
import time
from operator import itemgetter
import xlsxwriter
import os.path
from openpyxl import load_workbook
import openpyxl

def function():
    print('procedure')


def function2(x):
    print(x)

def float_str(value):
    if (type(value) == float):
        if value % 1 == 0:
            return str(value).split(".")[0]
        else:
            return str(float(value)).replace(".", ",")
    else:
        return str(value).strip()


def creat_level_tables(path):
    levels = poisk_level_tables(path)
    table = xlrd.open_workbook(path)

    df = []
    for i in range(len(table.sheet_names())):
        max_str = table.sheet_by_index(i).nrows
        max_levels = max(levels[i])
        df.append(Analis_Table_List(table, i, 0, max_str))
        level_str = [""] * (max_levels + 1)
        for j in range(len(df[i])):
            level = levels[i][j]
            if (level == 0):
                level_str = [""] * (max_levels + 1)
                level_str[0] = df[i][j][0]
                for k in range(1, max_levels + 1):
                    level_str[k] = df[i][j][0] + "(Уровень " + str(k) + ")"
            else:
                level_str = level_str[:level] + [df[i][j][0]] + [""] * (max_levels - level)
            df[i][j] = ["Уровень " + str(level)] + level_str + df[i][j][1:-2]
    return [df, table.sheet_names()]


def poisk_level_tables(path):
    table = load_workbook(path)
    levels = []
    for i in range(len(table.sheetnames)):
        sheet = table[table.sheetnames[i]]
        levels.append([])
        # max_group = 0
        for j in range(1, sheet.max_row + 1):
            levels[i].append(sheet.row_dimensions[j].outline_level)
    del table
    return levels
 

#Удаляет лишние столбцы не участвуюзие в  анализе
#формат str
def del_stb(table, stb, stb_selected):
    # print(stb,"//",stb_selected)
    # print(table[0:2])
    for i in range(len(table)):
        j = 0
        count_del = 0
        while j + count_del < len(stb):
            if(stb[j + count_del] not in stb_selected):
                count_del += 1
                del table[i][j]
            else:
                j += 1
    # print(table[0:2])
    # print("!!!!!")
    j = 0
    while j < len(stb):
        if(stb[j] not in stb_selected):
            del stb[j]
        else:
            j += 1
    # print(stb)


#Подсчитывает выбранные стб и удаляет лишние строки
def list_stb_calculations(table, add_stb, list_of_keys_selected_ind, list_of_keys_selected):

    for i in range(len(table)):
        if (table[i][-5] == "Да"):
            table[i] = table[i][:-6] + [" "] * len(list_of_keys_selected_ind) + table[i][-6:]
            continue

        for k in range(len(list_of_keys_selected_ind)):
            i2 = i
            x1 = 0
            while i2 < len(table) and table[i2][-2] == table[i][-2]:
                try:
                    if (table[i2][-5] != "Да"):
                        x1 += float(str(table[i2][list_of_keys_selected_ind[k]]).replace(",", "."))
                    # print(x1,"!!!!!!!!!!!!!")
                except BaseException:
                    x1 += 0
                if i2 > i and k + 1 == len(list_of_keys_selected_ind):
                    table[i2][-5] = "Да"
                i2 += 1
            table[i] = table[i][:-6] + [str(x1).replace(".", ",")] + table[i][-6:]
    for k in range(len(list_of_keys_selected_ind)):
        add_stb.append(list_of_keys_selected[k] + "(Сумма)")

# Возвращает лист названий столбцов
def Analis_Table_keys(table, index_list, start_id = 1):
    start_id = start_id - 1
    # print(table)
    stb = Analis_Table_List(table, index_list, start_id, start_id + 1)[0][:-2]
    # print(stb)
    flag = False
    for i in range(len(stb)):
        if(i == len(stb) - 1):
            if flag:
                stb[i] = str(i + 1) + "  " + stb[i]
        else:
            if stb[i] == stb[i + 1]:
                flag = True
                stb[i] = str(i + 1) + "  " + stb[i]
            elif flag:
                flag = False
                stb[i] = str(i + 1) + "  " + stb[i]
    return stb

def Auto_selection_start_str(table, index_list, start_id = 1):
    start_id = start_id - 1
    sheet = table.sheet_by_index(index_list)
    nrows = sheet.nrows
    i = (start_id + 1) % nrows
    # print(i)
    while(i != start_id):
        kkol = 0
        max = 1
        stb = Analis_Table_keys(table,  index_list, i + 1)
        for j in range(len(stb)):
            # print(stb)
            # print(stb[j])
            if(j >= 1):
                str1 = ")".join(stb[j].split(")")[1:])
                str2 = ")".join(stb[j - 1].split(")")[1:])
                if(stb[j] == stb[j-1]\
                        or len(str1) > 0 and str1 == str2):
                    max += 1
                    if (max >= len(stb) * 0.4):
                        # print("ril?")
                        kkol = 0
                        break
                else:
                    max = 1

            if(stb[j] != ""):
                try:
                    x = float(stb[j])
                    # x = str(stb[j]).split(".")[0]
                except BaseException:
                    kkol += 1
        if(kkol >= len(stb) * 0.8):
            return i + 1
        if(i >= 10000):
            return start_id + 1
        i = (i + 1) % nrows
    return i + 1
    # print(sheet.cell(0, 0).value)

# add "id"
# add "Удалить"
def Analis_Table_List(table, index_list, start_id = 1, exit_id = -1):
    all_data = []
    sheet = table.sheet_by_index(index_list)  # list.index
    crez = False
    if exit_id == -1:
        crez = True
        exit_id = sheet.nrows
    # Выгружаем данные
    for row_index in range(start_id, exit_id):
        row = []
        for col_index in range(0, sheet.ncols):
            if (str(sheet.cell(rowx=row_index, colx=col_index).ctype) == "5"):  # !!!!
                value = ""

            elif (str(sheet.cell(rowx=row_index, colx=col_index).ctype) == "2"):  # тип number
                if sheet.cell(rowx=row_index, colx=col_index).value % 1 == 0:
                    value = str(sheet.cell(rowx=row_index, colx=col_index).value).split(".")[0]
                else:
                    value = str(float(sheet.cell(rowx=row_index, colx=col_index).value)).replace(".", ",")


            elif (str(sheet.cell(rowx=row_index, colx=col_index).ctype) == "3"): # тип Date
                value = sheet.cell(rowx=row_index, colx=col_index).value
                y, m, d, h, i, s = xlrd.xldate_as_tuple(value, table.datemode)
                value = str("{0}.{1}.{2}".format(d, m, y))
            else:
                value = str(sheet.cell(rowx=row_index, colx=col_index).value).strip()
            row.append(value)

        row.append(row_index + 1)  # add "id"
        row.append("Нет")  # add Удалить"
        all_data.append(row)

    for crange in sheet.merged_cells:
        rlo, rhi, clo, chi = crange
        for rowx in range(rlo, rhi):
            for colx in range(clo, chi):
                if (rowx - start_id > -1 and rowx - start_id < len(all_data)):
                    if (str(sheet.cell(rlo, clo).ctype) == "5"): 
                        all_data[rowx - start_id][colx] = ""
                    elif (type(sheet.cell(rlo, clo).value) == float):
                        if sheet.cell(rlo, clo).value % 1 == 0:
                            # all_data[rowx - start_id][colx] = str(int(sheet.cell(rlo, clo).value))
                            all_data[rowx - start_id][colx] = str(sheet.cell(rlo, clo).value).split(".")[0]
                        else:
                            all_data[rowx - start_id][colx] = str(float(sheet.cell(rlo, clo).value)).replace(".", ",")
                    else:
                        all_data[rowx - start_id][colx] = str(sheet.cell(rlo, clo).value)
                # print(sheet.cell(rlo, clo).value, " " ,rlo, " " ,clo)

    #Острезаем снизу если пустые строки
    if crez:
        final = 0
        for i in range(1, len(all_data)):
            flag = False
            for j in range(0, sheet.ncols):
                if (all_data[i][j] != ""):
                    final = i
                    flag = True
                    break
            if(not flag and i >= 10):
                break
        return all_data[:final + 1]
    else:
        return all_data

    # print(all_data[:final + 1])
    return all_data[:final + 1]


# add - Добавить столбец id
def dict_list(df, add = False):
    # print("!")
    keys = list(df.keys())
    lenth = len(df[keys[0]])
    write_list = []

    item_format = {}

    for key in keys:
        item_format.update({key: ""})

    if (add):
        item_format.update({"id": ""})
        item_format.update({"Удалить": "Нет"})

    item = item_format
    for i in range(lenth):
        # if(i%1000 == 0):
        #     print('Загрузка : ' + str(int(i/lenth*100))+" %               ",  end='\r')

        for key in keys:
            item[key] = df[key][i]

        if (add):
            item["id"] = i

        write_list.append(dict(item))

    # print('Загрузка : ' + "100 %               ",  end='\r')

    return write_list


def list_dict(list):
    keys = list[0].keys()
    lenth = len(list)
    write_dict = {}
    for key in keys:
        write_dict.update({key: []})

    for i in range(lenth):
        # if(i%1000 == 0 ):
        #     print('Загрузка : ' + str(int(i/lenth*100))+" %               ",  end='\r')
        # print(i)

        for key in keys:
            # print(i)
            # print(key)
            # print(list[i][key])
            # print("_____________")
            write_dict[key].append(list[i][key])

    # print('Загрузка : ' +"100 %               ",  end='\r')

    return write_dict

# 2.0 Сортировка по одному из стб (и приведение типов)
# stb - int
# casting - приведение типов к int
def sort_list(list, stb, casting = False, reverse_flag = False):
    # Приведение типов
    if (casting):
        for i in range(len(list)):
            for j in stb:
                list[i][j] = int(list[i][j])

    # Сортировка по первому нужному стб
    list.sort(key=lambda x: x[stb[0]], reverse=reverse_flag)

    # Сортировка по остальным стб
    for i in range(1, len(stb)):
        i2 = 0
        while (i2 < len(list)):
            i3 = i2
            while (i3 < len(list) and list[i3][stb[i - 1]] == list[i2][stb[i - 1]]):
                i4 = i - 2
                flag = False
                while (i4 >= 0):
                    if (list[i3][stb[i4]] != list[i2][stb[i4]]):
                        flag = True
                        break
                    i4 = i4 - 1
                if (flag):
                    break
                i3 = i3 + 1
            # print("!!!")
            list[i2:i3] = sorted(list[i2:i3], key=lambda x: x[stb[i]], reverse=reverse_flag)
            # print("!!!2")
            i2 = i3


# add "Подгруппа"
# add "Колличество"
# add "Значние ключа"
# add "Расшифровка ключа"
# 2.0 Группировка и подсчет уникальных ключей (в уже отсортированном двумерном массиве)
# Удаленные, а потом нет
def group_list(list, stb, name_stb, flag_group_keys = False):
    namesSTB = str(name_stb).replace("[", "").replace("]", "").replace("'", "")
    for i in range(len(list)):
        KEY = ""
        for id in stb:
            if flag_group_keys:
                KEY = KEY + list[i][id]
            else:
                KEY = KEY + "+" + list[i][id]

        KEY = KEY.replace("[", "").replace("]", "").replace("'", "").replace(" ", "")
        KEY = KEY.upper()

        list[i] += ["", 0, KEY, namesSTB]

    sort_list(list, [len(list[0]) - 5])
    # print(list)

    i = 0
    id_del = len(list[0]) - 5
    group_id = len(list[0]) - 4
    while (i < len(list)):
        if (list[i][id_del] == "Да"):
            list[i][group_id] = -1
            i += 1
            continue
        break

    list[i:] = sorted(list[i:], key=lambda x: len(list[0]) - 2)

    group = 1
    kkol_id = len(list[0]) - 3
    key_id = len(list[0]) - 2
    n = len(list)
    while (i < n):
        i2 = i + 1
        while (i2 < n and list[i][key_id] == list[i2][key_id]):
            i2 += 1
        kkol = i2 - i
        while (i < i2):
            list[i][group_id] = group
            list[i][kkol_id] = kkol
            i += 1
        group += 1


# 2.0 stb - 0 список выбраных фильтров 1-3 список ключей
def filter(list, stb, param, pos_stb = []):
    # Примение фильтра 2
    if (param == 0):
        for i in range(len(list)):
            for st in stb.keys():
                ind_stb = pos_stb.index(st)
                if (len(stb[st]) > 0 and str(list[i][ind_stb]) not in stb[st]):
                    list[i][-1] = "Да"
                    # print("!3")
                    break

    # Примение фильтра 1 пункт 2
    elif (param == 1):
        group = 1
        max1 = 0
        max2 = 0
        id_max = -1
        Len_list = len(list)
        len_str = len(list[0])
        for i in range(Len_list):
            if (list[i][len_str - 4] == -1):
                id_max = i + 1
                continue

            if (i == Len_list - 1 and id_max >= 0):
                list[id_max][len_str - 5] = "Нет"

            if (list[i][len_str - 3] < 2):
                continue

            if (list[i][len_str - 4] != group):
                if (id_max != -1):
                    list[id_max][len_str - 5] = "Нет"
                id_max = i
                group = list[i][len_str - 4]

                max1 = 0
                max2 = 0

            list[i][len_str - 5] = "Да"
            kkol1 = 0
            kkol2 = 0

            for j in range(len_str - 6):
                try:
                    if (str(list[i][j]) == "nan" or str(list[i][j]) == ""):
                        x = 1 / 0
                    kkol2 = kkol2 + 1
                    if (j in stb):
                        kkol1 = kkol1 + 1
                except:
                    pass

            if (kkol1 > max1):
                id_max = i
                max1 = kkol1
                max2 = kkol2
            elif (kkol1 == max1):
                if (kkol2 >= max2):
                    id_max = i
                    max1 = kkol1
                    max2 = kkol2

            if (i == len(list) - 1):
                list[id_max][len_str - 5] = "Нет"

            # print(kkol1," ",kkol2," ",i," ",group," ",max1," ",max2," ",id_max)

    # Слияние ключей
    elif (param == 2):
        len_str = len(list[0])
        i = 0
        n = len(list)
        while (i < n):
            if (list[i][len_str - 4] == -1):
                id_max = i + 1
                continue
            i2 = i + 1
            list[i][len_str - 5] = "Нет"
            while (i2 < n and list[i][len_str - 4] == list[i2][len_str - 4]):
                for j in range(len_str - 6):
                    if (str(list[i][j]) == "nan" or str(list[i][j]) == ""):
                        if (str(list[i2][j]) != "nan" and str(list[i2][j]) != ""):
                            list[i][j] = str(list[i2][j])
                list[i2][len_str - 5] = "Да"
                i2 = i2 + 1
            i = i2

def write_table_defolt(table = [], names_list = [], file_name = "", apdate_file = False):
    path = os.getcwd() + "\\Промежуточные файлы\\"
    try:
        os.mkdir(path)
    except OSError:
        pass

    if(apdate_file):
        i = 2
        while (os.path.exists(path + file_name + ".xlsx")):
            if i == 2:
                file_name = file_name + "(2)"
            else:
                file_name = "(".join(file_name.split("(")[:-1]) + "(" + str(i) + ")"
            i += 1
    workbook = xlsxwriter.Workbook(path + file_name + ".xlsx")
    normal_format = workbook.add_format({"size": 12,
                                         "align": "center", "border_color": "#000000", 'border': 1})
    for i in range(len(names_list)):
        try:
            worksheet = workbook.add_worksheet(names_list[i])
        except:
            worksheet = workbook.add_worksheet(str(i + 1))
        worksheet.set_column(0, len(table[i][0]), 15)  # Ширина
        for j in range(len(table[i])):
            for k in range(len(table[i][j])):
                worksheet.write(j, k, str(table[i][j][k]), normal_format)
    workbook.close()
    return path + file_name + ".xlsx"



# 2.0
#keys - все наши ключи, key + add_stb(str)
#keys_pos - выбранные ключи (уникальный ключ)

def write_table(list, keys, file_name, params, color = False, add_stb = 0, keys_pos = []):
    i = 2
    Obrabit = "(Обработанный)"
    path = os.getcwd() + "\\Результаты\\"
    try:
        os.mkdir(path)
    except OSError:
        pass
    while (os.path.exists(path + file_name + ".xlsx")):
        # print(file_name)
        file_name = Obrabit.join(file_name.split(Obrabit)[:-1]) + Obrabit + "(" + str(i) + ")"
        i += 1

    workbook = xlsxwriter.Workbook(path + file_name + ".xlsx")
    stb_format = workbook.add_format({'bold': True, "size": 12,
                                      "align": "center", "border_color": "#000000", 'border': 5})
    del_format = workbook.add_format({'bg_color': '#A6A6A6', "size": 12,
                                      "align": "center", "border_color": "#000000", 'border': 1})
    normal_format = workbook.add_format({"size": 12,
                                         "align": "center", "border_color": "#000000", 'border': 1})
    # number_format = workbook.add_format({'num_format': '#,##0.00', "size": 12,
    #                                   "align": "center", "border_color": "#000000", 'border': 1})

    stb_green_2_format = workbook.add_format(
        {'bold': True, "size": 12, "align": "center", 'bg_color': "#5FD4B1", 'border': 5})
    stb_pink_format = workbook.add_format(
        {'bold': True, "size": 12, "align": "center", 'bg_color': "#E6B8B7", 'border': 5})
    stb_green_format = workbook.add_format(
        {'bold': True, "size": 12, "align": "center", 'bg_color': "#D8E4BC", 'border': 5})
    stb_blue_format = workbook.add_format(
        {'bold': True, "size": 12, "align": "center", 'bg_color': "#8DB4E2", 'border': 5})
    stb_yellow_format = workbook.add_format(
        {'bold': True, "size": 12, "align": "center", 'bg_color': "#ffff00", 'border': 5})
    stb_yellow_2_format = workbook.add_format(
        {'bold': True, "size": 12, "align": "center", 'bg_color': "#FFC000", 'border': 5})
    stb_yellow_3_format = workbook.add_format(
        {'bold': True, "size": 12, "align": "center", 'bg_color': "#E26B0A", 'border': 5})

    #Полное совпадение
    green_2_format = workbook.add_format({"size": 12, "align": "center", 'bg_color': "#67EBAC", 'border': 1})
    pink_format = workbook.add_format({"size": 12, "align": "center", 'bg_color': "#ed1581", 'border': 1})
    green_format = workbook.add_format({"size": 12, "align": "center", 'bg_color': "#0ceb22", 'border': 1})
    blue_format = workbook.add_format({"size": 12, "align": "center", 'bg_color': "#2846de", 'border': 1})

    num_format = workbook.add_format({'num_format': "### ### ### ### ### ### ### ### ### ##0.00", 'bold': True})
    worksheet = workbook.add_worksheet("Результат")
    len_list = len(list)
    len_stb = len(keys)


    for i in range(len_stb - add_stb):
        if (color and i in keys_pos):
            worksheet.write(0, i, str(keys[i]), stb_blue_format)
        else:
            worksheet.write(0, i, str(keys[i]), stb_format)

    calcul_stb = set()
    for i in range(len_stb - add_stb, len_stb):
        if i in calcul_stb:
            continue
        if (color):
            if ("(Элемент 1)" in str(keys[i])):
                calcul_stb.add(i)
                worksheet.write(0, i, str(keys[i]), stb_yellow_format)
            elif ("(Элемент 2)" in str(keys[i])):
                calcul_stb.add(i)
                worksheet.write(0, i, str(keys[i]), stb_yellow_2_format)
            elif ("(Разность)" in str(keys[i])):
                calcul_stb.add(i)
                worksheet.write(0, i, str(keys[i]), stb_yellow_3_format)
            else:
                worksheet.write(0, i, str(keys[i]), stb_green_2_format)
        else:
            worksheet.write(0, i, str(keys[i]), stb_format)
    if (color):
        worksheet.write(0, len_stb, "id", stb_pink_format)
        worksheet.write(0, len_stb + 1, "Удалить", stb_pink_format)
        worksheet.write(0, len_stb + 2, "Подгруппа", stb_green_format)
        worksheet.write(0, len_stb + 3, "Колличество", stb_green_format)
        worksheet.write(0, len_stb + 4, "Составленный ключ", stb_green_format)
        worksheet.write(0, len_stb + 5, "Расшифровка ключа", stb_green_format)

    else:
        worksheet.write(0, len_stb, "id", stb_format)
        worksheet.write(0, len_stb + 1, "Удалить", stb_format)
        worksheet.write(0, len_stb + 2, "Подгруппа", stb_format)
        worksheet.write(0, len_stb + 3, "Колличество", stb_format)
        worksheet.write(0, len_stb + 4, "Составленный ключ", stb_format)
        worksheet.write(0, len_stb + 5, "Расшифровка ключа", stb_format)

    worksheet.set_column(0, len_stb + 4, 15)  # Ширина столбца
    worksheet.set_column(len_stb + 4, len_stb + 5, 21)  # Ширина столбца

    ind_del = 0
    for i in range(len_list):
        Format = normal_format
        if (str(list[i][len_stb + 1]) == "Да"):
            if (params == 1):
                Format = del_format
            elif (params == 2):
                ind_del -= 1
                continue
        elif (color):
            if (str(list[i][len_stb - 1]) == "Полное"):
                Format = green_2_format
        for j in range(len_stb):
            worksheet.write(i + 1 + ind_del, j, str(list[i][j]), Format)
        for j in calcul_stb:
            try:
                worksheet.write(i + 1 + ind_del, j, float(list[i][j].replace(",", ".")), num_format)
            except BaseException:
                worksheet.write(i + 1 + ind_del, j, 0, num_format)

        worksheet.write(i + 1 + ind_del, len_stb, list[i][len_stb], Format)
        worksheet.write(i + 1 + ind_del, len_stb + 1, str(list[i][len_stb + 1]), Format)
        worksheet.write(i + 1 + ind_del, len_stb + 2, list[i][len_stb + 2], Format)
        worksheet.write(i + 1 + ind_del, len_stb + 3, list[i][len_stb + 3], Format)
        worksheet.write(i + 1 + ind_del, len_stb + 4, str(list[i][len_stb + 4]), Format)
        worksheet.write(i + 1 + ind_del, len_stb + 5, str(list[i][len_stb + 5]), Format)

    workbook.close()
    return file_name + ".xlsx"

